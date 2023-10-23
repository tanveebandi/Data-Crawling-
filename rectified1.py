import requests
from bs4 import BeautifulSoup
import openpyxl
import os
from textblob import TextBlob
import textstat  # Import the textstat library for text analysis

# Specify the path to your Excel input file
input_file_path = r'C:\Users\gajen\Downloads\Input.xlsx'

# Specify the path to your output Excel file
output_file_path = r'C:\Users\gajen\Downloads\Output Data Structure.xlsx'

# Create a directory to save the extracted articles
output_dir = r'C:\Users\gajen\Downloads\Extracted_Articles'
os.makedirs(output_dir, exist_ok=True)

# Open the input Excel file
workbook = openpyxl.load_workbook(input_file_path)
worksheet = workbook.active

# Create a new Excel workbook for output
output_workbook = openpyxl.Workbook()
output_worksheet = output_workbook.active

# Add column headers to the output worksheet
output_worksheet.append([
    "URL_ID",
    "Sentiment_Polarity",
    "Sentiment_Subjectivity",
    "FOG_Index",
    "Complex_Word_Count",
    "Word_Count",
    "Syllables_Per_Word",
    "Personal_Pronouns",
    "Avg_Word_Length",
    "Avg_Sentence_Length",
    "Avg_Words_Per_Sentence"
])

# Iterate through the rows in the input worksheet and extract data
for row in worksheet.iter_rows(values_only=True):
    url_id = row[0]
    url = row[1]

    # Check if the URL is valid
    if not url.startswith("https"):
        print(f"Invalid URL: {url}")
        continue

    # Send an HTTP request to the URL
    response = requests.get(url)

    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(response.text, 'html.parser')

        # Extract the text from the HTML
        text = soup.get_text()
        title = soup.find('h1').text if soup.find('h1') else "Title not found"#for extracting the title of the article 
        content = "\n".join(p.text for p in soup.find_all('p'))#for extracting the content of article 

        # Perform sentiment analysis using TextBlob library 
        blob = TextBlob(text)
        sentiment = blob.sentiment
        sentiment_polarity = sentiment.polarity
        sentiment_subjectivity = sentiment.subjectivity

        # Calculate FOG index using textstat library
        fog_index = textstat.gunning_fog(text)

        # Calculate complex word count
        def calculate_complex_word_count(text):
            complex_word_count = sum(1 for word in text.split() if textstat.syllable_count(word) > 3)
            return complex_word_count

        complex_word_count = calculate_complex_word_count(text)

        # Calculate personal pronouns count
        def calculate_personal_pronouns(text):
            personal_pronouns = ["I", "me", "my", "mine", "myself", "you", "your", "yours", "yourself", "he", "him", "his", "himself", "she", "her", "hers", "herself", "it", "its", "itself", "we", "us", "our", "ours", "ourselves", "they", "them", "their", "theirs", "themselves"]
            personal_pronoun_count = sum(1 for word in text.split() if word.lower() in personal_pronouns)
            return personal_pronoun_count

        personal_pronouns = calculate_personal_pronouns(text)

        # Calculate average word length
        avg_word_length = textstat.avg_letter_per_word(text)

        # Calculate average sentence length. Here we are initiatinga function of our own. 
        def calculate_avg_sentence_length(text):
            sentences = text.split('.')
            words = text.split()
            average_length = len(words) / len(sentences)
            return average_length

        avg_sentence_length = calculate_avg_sentence_length(text)

        # Calculate average words per sentence. Here we are initiating a function of our own. 
        def calculate_avg_words_per_sentence(text):
            sentences = text.split('.')
            words = text.split()
            avg_words = len(words) / len(sentences)
            return avg_words

        avg_words_per_sentence = calculate_avg_words_per_sentence(text)

        # Append the data to the output worksheet
        output_worksheet.append([
            url_id,
            sentiment_polarity,
            sentiment_subjectivity,
            fog_index,
            complex_word_count,
            textstat.lexicon_count(text, removepunct=True),
            textstat.syllable_count(text),
            personal_pronouns,
            avg_word_length,
            avg_sentence_length,
            avg_words_per_sentence
        ])

        # Use the URL_ID for the file name and save the article text
        article_file_path = os.path.join(output_dir, f"{url_id}.txt")
        with open(article_file_path, 'w', encoding='utf-8') as article_file:
            article_file.write(f"Title: {title}\n\n")
            article_file.write(f"Content:\n{content}\n\n")
            article_file.write(f"Full Text:\n{text}")

        print(f"Article extracted and saved: {article_file_path}")

    else:
        print(f"Failed to retrieve data from URL: {url}. Status code: {response.status_code}")

# Save the output workbook to the specified output file
output_workbook.save(output_file_path)


workbook.close()
output_workbook.close()
